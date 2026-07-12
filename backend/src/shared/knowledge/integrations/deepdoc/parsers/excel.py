from __future__ import annotations

from io import BytesIO
import logging
import re

import pandas as pd

from novamind.shared.knowledge.integrations.deepdoc.compat import LazyImage, find_codec

# Adapted from RAGFlow deepdoc/parser/excel_parser.py

ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


class RAGFlowExcelParser:
    @staticmethod
    def _import_openpyxl():
        from openpyxl import Workbook, load_workbook

        return Workbook, load_workbook

    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        _, load_workbook = RAGFlowExcelParser._import_openpyxl()
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        if not (file_head.startswith(b"PK\x03\x04") or file_head.startswith(b"\xd0\xcf\x11\xe0")):
            logging.info("DeepDoc excel parser received non-Excel payload; attempting CSV fallback")
            df = pd.read_csv(file_like_object, on_bad_lines="skip")
            return RAGFlowExcelParser._dataframe_to_workbook(df)

        try:
            return load_workbook(file_like_object, data_only=True)
        except Exception as exc:
            logging.info("openpyxl load failed, trying pandas fallback: %s", exc)
            file_like_object.seek(0)
            try:
                dfs = pd.read_excel(file_like_object, sheet_name=None)
            except Exception:
                file_like_object.seek(0)
                dfs = pd.read_excel(file_like_object, engine="calamine")
            return RAGFlowExcelParser._dataframe_to_workbook(dfs)

    @staticmethod
    def _clean_dataframe(df: pd.DataFrame):
        def clean_string(value):
            if isinstance(value, str):
                return ILLEGAL_CHARACTERS_RE.sub(" ", value)
            return value

        return df.apply(lambda col: col.map(clean_string))

    @staticmethod
    def _fill_worksheet_from_dataframe(ws, df: pd.DataFrame):
        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

    @staticmethod
    def _dataframe_to_workbook(df):
        Workbook, _ = RAGFlowExcelParser._import_openpyxl()
        if isinstance(df, dict) and len(df) > 1:
            return RAGFlowExcelParser._dataframes_to_workbook(df)

        df = RAGFlowExcelParser._clean_dataframe(df)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        RAGFlowExcelParser._fill_worksheet_from_dataframe(ws, df)
        return wb

    @staticmethod
    def _dataframes_to_workbook(dfs: dict):
        Workbook, _ = RAGFlowExcelParser._import_openpyxl()
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_name, df in dfs.items():
            df = RAGFlowExcelParser._clean_dataframe(df)
            ws = wb.create_sheet(title=sheet_name)
            RAGFlowExcelParser._fill_worksheet_from_dataframe(ws, df)
        return wb

    @staticmethod
    def _extract_images_from_worksheet(ws, sheetname=None):
        images = getattr(ws, "_images", [])
        if not images:
            return []

        raw_items = []
        for img in images:
            try:
                img_bytes = img._data()
                lazy_img = LazyImage([img_bytes])
                anchor = img.anchor
                if hasattr(anchor, "_from") and hasattr(anchor, "_to"):
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = anchor._to.row + 1, anchor._to.col + 1
                    span = "single_cell" if (r1 == r2 and c1 == c2) else "multi_cell"
                else:
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = r1, c1
                    span = "single_cell"
                raw_items.append(
                    {
                        "sheet": sheetname or ws.title,
                        "image": lazy_img,
                        "image_description": "",
                        "row_from": r1,
                        "col_from": c1,
                        "row_to": r2,
                        "col_to": c2,
                        "span_type": span,
                    }
                )
            except Exception:
                continue
        return raw_items

    @staticmethod
    def _get_actual_row_count(ws):
        max_row = ws.max_row
        if not max_row:
            return 0
        if max_row <= 10000:
            return max_row

        max_col = min(ws.max_column or 1, 50)

        def row_has_data(row_idx):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    return True
            return False

        if not any(row_has_data(i) for i in range(1, min(101, max_row + 1))):
            return 0

        left, right = 1, max_row
        last_data_row = 1
        while left <= right:
            mid = (left + right) // 2
            found = False
            for row_idx in range(mid, min(mid + 10, max_row + 1)):
                if row_has_data(row_idx):
                    found = True
                    last_data_row = max(last_data_row, row_idx)
                    break
            if found:
                left = mid + 1
            else:
                right = mid - 1

        for row_idx in range(last_data_row, min(last_data_row + 500, max_row + 1)):
            if row_has_data(row_idx):
                last_data_row = row_idx
        return last_data_row

    @staticmethod
    def _get_rows_limited(ws):
        actual_rows = RAGFlowExcelParser._get_actual_row_count(ws)
        if actual_rows == 0:
            return []
        return list(ws.iter_rows(min_row=1, max_row=actual_rows))

    def html(self, fnm, chunk_rows=256):
        from html import escape

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        tb_chunks = []

        def fmt(value):
            if value is None:
                return ""
            return str(value).strip()

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = RAGFlowExcelParser._get_rows_limited(ws)
            if not rows:
                continue

            header_row = "<tr>" + "".join(f"<th>{escape(fmt(cell.value))}</th>" for cell in list(rows[0])) + "</tr>"
            n_data_rows = len(rows) - 1
            for chunk_i in range((n_data_rows + chunk_rows - 1) // chunk_rows):
                table_html = f"<table><caption>{sheetname}</caption>{header_row}"
                for row in list(rows[1 + chunk_i * chunk_rows : min(1 + (chunk_i + 1) * chunk_rows, len(rows))]):
                    table_html += "<tr>"
                    for cell in row:
                        table_html += f"<td>{escape(fmt(cell.value))}</td>" if cell.value is not None else "<td></td>"
                    table_html += "</tr>"
                table_html += "</table>\n"
                tb_chunks.append(table_html)
        return tb_chunks

    def markdown(self, fnm):
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        try:
            if not isinstance(file_like_object, str):
                file_like_object.seek(0)
            df = pd.read_excel(file_like_object)
        except Exception:
            if isinstance(file_like_object, str):
                with open(file_like_object, "rb") as handle:
                    binary = handle.read()
            else:
                file_like_object.seek(0)
                binary = file_like_object.read()
            text = binary.decode(find_codec(binary), errors="ignore")
            df = pd.read_csv(BytesIO(text.encode("utf-8")), on_bad_lines="skip")
        df = df.replace(r"^\s*$", "", regex=True)
        return df.to_markdown(index=False)

    def __call__(self, fnm):
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)

        rows_out = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = RAGFlowExcelParser._get_rows_limited(ws)
            if not rows:
                continue
            header = list(rows[0])
            for row in list(rows[1:]):
                fields = []
                for idx, cell in enumerate(row):
                    if cell.value is None or str(cell.value).strip() == "":
                        continue
                    title = str(header[idx].value) if idx < len(header) else ""
                    prefix = f"{title}: " if title else ""
                    fields.append(prefix + str(cell.value))
                if not fields:
                    continue
                line = "; ".join(fields)
                if "sheet" not in sheetname.lower():
                    line += f" -- {sheetname}"
                rows_out.append(line)
        return rows_out

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += RAGFlowExcelParser._get_actual_row_count(ws)
            return total

        if fnm.split(".")[-1].lower() in {"csv", "txt"}:
            encoding = find_codec(binary)
            return len(binary.decode(encoding, errors="ignore").split("\n"))
        return 0
